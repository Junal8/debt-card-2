#!/usr/bin/env python
# coding: utf-8

# In[1]:


import pandas as pd
#from collections import Counter
import numpy as np
#from math import factorial
from IPython.display import display  
import openpyxl as op
import shutil
import os
from datetime import datetime
from openpyxl import Workbook
from pandas import ExcelWriter
import glob
from PIL import Image
from openpyxl.styles import PatternFill
pd.set_option('display.max_columns', None)
#from openpyxl.styles import PatternFill
#from openpyxl.styles.differential import DifferentialStyle
#from openpyxl.formatting.rule import Rule


# # ✅Выгрузка пункт 1

# In[2]:


#Считываем за прошлый год файл пункт 1
data_history = pd.read_excel(r"Z:\Общая папка МИ по УД\Долговые карты субъектов\пункты на 01.09.2022\пункт 1\01.01.2021-01.01.2022.xlsx", 
                             header =5,
                  converters={'Unnamed: 0' : str,                              
                             },
                  #sep=';'
                  #dtype=str
                 )


#Считываем за нынешний год файл пункт 1
data_new = pd.read_excel(r"Z:\Общая папка МИ по УД\Долговые карты субъектов\пункты на 01.09.2022\пункт 1\01.01.2022-01.09.2022.xlsx", 
                             header =5,
                  converters={'Unnamed: 0' : str,                              
                             },
                  #sep=';'
                  #dtype=str
                 )

# заполнение шапки с датам
data_history_head = op.load_workbook(r"Z:\Общая папка МИ по УД\Долговые карты субъектов\пункты на 01.09.2022\пункт 1\01.01.2021-01.01.2022.xlsx", 
                                     data_only=True).active
data_new_head = op.load_workbook(r"Z:\Общая папка МИ по УД\Долговые карты субъектов\пункты на 01.09.2022\пункт 1\01.01.2022-01.09.2022.xlsx", 
                                     data_only=True).active

head_c11 = data_history_head['C3'].value
head_g11 = data_history_head['G3'].value
head_k11 = data_history_head['K3'].value

head_c16 = data_new_head['C3'].value
head_g16 = data_new_head['G3'].value
head_k16 = data_new_head['K3'].value


# In[3]:


data_history


# # ✅ Выгрузка пункт 2

# In[4]:


#Считываем за нынешний год файл пункт 2
data_new_2 = pd.read_excel(r"Z:\Общая папка МИ по УД\Долговые карты субъектов\пункты на 01.09.2022\пункт 2\пункт 2.xlsx", 
                             header =4,
                  converters={'Unnamed: 3' : str,                              
                            },
                  #sep=';'
                  #dtype=str
                 )
# заполнение шапки с датам

data_new_2_head = op.load_workbook(r"Z:\Общая папка МИ по УД\Долговые карты субъектов\пункты на 01.09.2022\пункт 2\пункт 2.xlsx", 
                                     data_only=True).active

head_c21 = data_new_2_head['E2'].value
head_c21 = head_c21.strftime('%d.%m.%Y')
head_e21 = data_new_2_head['G2'].value
head_e21 = head_e21.strftime('%d.%m.%Y')


# # ✅ Выгрузка пункт 3

# In[5]:


#Считываем за прошлый год файл пункт 3
data_history_3 = pd.read_excel(r"Z:\Общая папка МИ по УД\Долговые карты субъектов\пункты на 01.09.2022\пункт 3\01.01.2022.xlsx", 
                             header = 7,
                  converters={'Unnamed: 0' : str,                              
                             },
                  #sep=';'
                  #dtype=str
                 )
#Считываем за нынешний год файл пункт 3
data_new_3 = pd.read_excel(r"Z:\Общая папка МИ по УД\Долговые карты субъектов\пункты на 01.09.2022\пункт 3\01.09.2022.xlsx", 
                             header = 7,
                  converters={'Unnamed: 0' : str,                              
                             },
                  #sep=';'
                  #dtype=str
                 )

regions = pd.read_excel(r"Z:\8-Отдел анализа данных\001 - Сотрудники\05 - Чурахин А.А\Dolg_karta.xlsx",                                               
                  dtype=str
                 )

# заполнение шапки с датам
data_history_3_head = op.load_workbook(r"Z:\Общая папка МИ по УД\Долговые карты субъектов\пункты на 01.09.2022\пункт 3\01.01.2022.xlsx", 
                                     data_only=True).active
data_new_3_head = op.load_workbook(r"Z:\Общая папка МИ по УД\Долговые карты субъектов\пункты на 01.09.2022\пункт 3\01.09.2022.xlsx", 
                                     data_only=True).active

head_c30 = data_history_3_head['B7'].value
head_f30 = data_new_3_head['B7'].value


# # ✅ Выгрузка пункт 4

# In[6]:


#Считываем за нынешний год файл пункт 2
data_new_4 = pd.read_excel(r"Z:\Общая папка МИ по УД\Долговые карты субъектов\пункты на 01.09.2022\пункт 4\Приложение 3_данные на 01.09.2022.xlsx", 
                             header =4,
                  converters={1 : str,
                              2 : str, 
                            },
                  #sep=';'
                  #dtype=str
                 )

data_new_4.rename(columns = { 1 : 'numb',
                        2 : 'regions', 
                       3 : 'col_3',
                       4 : 'col_4',
                       5 : 'col_5',
                       6 : 'col_6',
                       7 : 'col_7',
                       8 : 'col_8', 
                       9 : 'col_9',
                       10 : 'col_10',
                       11 : 'col_11', 
                       12 : 'col_12',
                       13 : 'col_13'                              
                             }, inplace = True)
# заполнение шапки с датам

#data_new_4_head = op.load_workbook("Z:\\Долговые карты субъектов\\пункты на 01.04.2022\\пункт 4\\Приложение №2 на 01.04.2022.xlsx", 
                                     #data_only=True).active


# # ✅ Выгрузка пункт 5

# In[7]:


#Считываем за нынешний год файл
data_new_5 = pd.read_excel(r"Z:\Общая папка МИ по УД\Долговые карты субъектов\пункты на 01.09.2022\пункт 5\Рейтинг 2022 08 01.xlsx", 
                             header = 9, 
                           nrows = 84,
                           names=['kod', 'reg', 'numb', '1.', '2.', '3.', '4.', '5.', '6.', '7.', '8.', '9.', '10.', '11.', 
                                  '12.', '13.', '14.', '15.', '16.', '17.', '18.', '19.', '20.', '21.', '22.', '23.', 
                                  '24.', '25.', '26.', '27.', '28.', '29.', '30.', '31.', '32.', '33.', '34.', '35.', 
                                  '36.', '37.', '38.', '39.', '40.', '41.', '42.', '43.', '44.', '45.', '46.', '47.', 
                                  '48.', '49.', '50.', '51.', '52.', '53.', '54.', '55.', '56.', '57.', '58.', '59.', 
                                  '60.', '61.', '62.', '63.', '64.', '65.', '66.', '67.', '68.', '69.',
                                 '70.', '71.', '72.', '73.', '74.', '75.'],
                           index_col=False,
                           converters={'reg' : str}
                              
                  #converters={'Unnamed: 3' : str,                              
                     #       },
                  #sep=';'
                  #dtype=str
                 )


# In[8]:


data_new_5.tail(5)


# In[9]:


# заполнение шапки с датам
data_new_5_head = op.load_workbook(r"Z:\Общая папка МИ по УД\Долговые карты субъектов\пункты на 01.09.2022\пункт 5\Рейтинг 2022 08 01.xlsx", 
                                     data_only=True).active

head_I38 = data_new_5_head['A2'].value
head_I38 = head_I38[24:46]


# In[10]:


# отбираем только те столбцы, что нам нужны
data_new_5 = data_new_5[['reg', 'numb', '8.', '9.', '14.', '15.', '23.', '26.', '29.', '33.', '36.', '39.', 
                        '40.', '47.', '50.',  '51.', '57.', '60.', '63.', '64.', '72.', '75.']]


# In[11]:


path_5 = r'Z:\Общая папка МИ по УД\Долговые карты субъектов\002 _ Данные _\01 _ Исходные сведения (не утв)\05\ОПТГ'
all_files_5 = glob.glob(path_5 + "/*.xlsx")


# In[12]:


data_5_left = []
for filename in all_files_5:
    df = pd.read_excel(filename, 
                     #sep=';', 
                     #index_col=None, 
                     header=9,
                    nrows = 85,
                     dtype=str,
                       usecols = "A:C",
                       names=['first', 'second', 'third']
                    )
    data_5_left.append(df)


# In[13]:


# поменялись таблицы и с 14 добавили код НО, в этих 2 ячейках привожу все к одному виду
def data_5_left_all(dt, n):
    dt[n] = dt[n].drop(columns=['third']#, inplace=True
                )
    dt[n] = dt[n].rename(columns = { 'first' : 'regions',
                          'second' : 'number'
                        }, 
              #inplace = True
             )
    
for i in range(14):
    data_5_left_all(data_5_left, i)


# In[14]:


#
def data_5_left_all(dt, n):
    dt[n] = dt[n].drop(columns=['first']#, inplace=True
                )
    dt[n] = dt[n].rename(columns = { 'second' : 'regions',
                          'third' : 'number'
                        }, 
              #inplace = True
             )
    
for i in range(14, len(data_5_left)):
    data_5_left_all(data_5_left, i)


# In[15]:


data_5_left


# # ✅ Выгрузка пункт 6

# In[16]:


#Считываем за прошлый год файл пункт 6
data_history_6 = pd.read_excel(r"Z:\Общая папка МИ по УД\Долговые карты субъектов\пункты на 01.09.2022\пункт 6\01.01.2022.xlsx", 
                             header =5,
                  #converters={'Unnamed: 0' : str,                              
                             #},
                  names=['reg', 'region', '1.', '2.', '3.', '4.', '5.', '6.', '7.', '8.', '9.', '10.', '11.', 
                         '12.', '13.', '14.', '15.', '16.', '17.', '18.', '19.', '20.', '21.', '22.', '23.', '24.'],
                                converters={'reg' : str,                              
                             },
                  #sep=';'
                  #dtype=str
                 )
#Считываем за нынешний год файл пункт 6
data_new_6 = pd.read_excel(r"Z:\Общая папка МИ по УД\Долговые карты субъектов\пункты на 01.09.2022\пункт 6\01.09.2022.xlsx", 
                             header =7,
                  #converters={'Unnamed: 0' : str,                              
                            # },
                  names=['reg', 'region', '1.', '2.', '3.', '4.', '5.', '6.', '7.', '8.', '9.', '10.', '11.', 
                         '12.', '13.', '14.', '15.', '16.', '17.', '18.', '19.', '20.', '21.', '22.', '23.', '24.'] ,
                           converters={'reg' : str,                              
                             },
                  index_col=False,
                  #sep=';'
                  #dtype=str
                 )


# In[17]:


data_new_6.head(4)


# In[18]:


# заполнение шапки с датам
data_history_6_head = op.load_workbook(r"Z:\Общая папка МИ по УД\Долговые карты субъектов\пункты на 01.09.2022\пункт 6\01.01.2022.xlsx", 
                                     data_only=True).active
data_new_6_head = op.load_workbook(r"Z:\Общая папка МИ по УД\Долговые карты субъектов\пункты на 01.09.2022\пункт 6\01.09.2022.xlsx", 
                                     data_only=True).active

head_a55 = data_history_6_head['H2'].value
head_a55 = head_a55[13:32] + '.'
head_a61 = data_new_6_head['I3'].value


# # ✅ Выгрузка пункт 7

# In[19]:


path_7 = r'Z:\Общая папка МИ по УД\Долговые карты субъектов\пункты на 01.09.2022\пункт 7'
all_files_7 = glob.glob(path_7 + "/*/Ф7.xls")


# In[20]:


data_7 = []
for filename in all_files_7:
    df = pd.read_excel(filename,
                       sheet_name='Справочно к Разделам I,II (3)',
                     #sep=';', 
                     #index_col=None, 
                     header=4,
                     nrows = 6,
                     dtype=str,
                     usecols = "A:F",
                     names=['type', 'ne_nygno', 'all', 'UL', 'FL', 'IP']
                    )
    data_7.append(df)


# In[21]:


# без КН
list_region_7 = []
for i in range(len(all_files_7)-9):
    list_region_7.append(all_files_7[i][78:80]) 
    
for i in range(len(all_files_7)-9, len(all_files_7)):
    list_region_7.append(all_files_7[i][78:82])     


# # ✅ Выгрузка пункт 8

# In[22]:


#Считываем за прошлый год файл пункт 8
data_history_8 = pd.read_excel(r"Z:\Общая папка МИ по УД\Долговые карты субъектов\пункты на 01.09.2022\пункт 8\01.09.2021.xlsx", 
                               sheet_name='1190',
                               header =9,
                               names=['reg', '1.', '2.', '3.', '4.'],
                  converters={'Unnamed: 0' : str,                              
                             },
                  #sep=';'
                  #dtype=str
                 )
#Считываем за нынешний год файл пункт 8
data_new_8 = pd.read_excel(r"Z:\Общая папка МИ по УД\Долговые карты субъектов\пункты на 01.09.2022\пункт 8\01.09.2022.xlsx", 
                             header = 9,
                             sheet_name='1190',
                             names=['reg', '1.', '2.', '3.', '4.'],
                  converters={'reg' : str,                              
                             },
                  #sep=';'
                  #dtype=str
                 )


# In[23]:


# заполнение шапки с датам
data_history_8_head = op.load_workbook(r"Z:\Общая папка МИ по УД\Долговые карты субъектов\пункты на 01.09.2022\пункт 8\01.09.2021.xlsx", 
                                     data_only=True).active
data_new_8_head = op.load_workbook(r"Z:\Общая папка МИ по УД\Долговые карты субъектов\пункты на 01.09.2022\пункт 8\01.09.2022.xlsx", 
                                     data_only=True).active


# In[24]:


head_i69 = data_history_8_head['A4'].value
head_i69 = head_i69[16:28] + '.'
head_g69 = data_new_8_head['A4'].value
head_g69 = head_g69[16:28] + '.'


# # ✅ Выгрузка пункт 9

# In[25]:


#Считываем за прошлый год файл пункт 9
data_9 = pd.read_excel(r"Z:\Общая папка МИ по УД\Долговые карты субъектов\пункты на 01.09.2022\пункт 9\Расчет на 01.09.2022.xlsx", 
                               #sheet_name='1190',
                               header = 3,
                               names=['reg', 'region', '69.', '46.', '47.', 'Kytp.'],
                               converters={'reg' : str,                              
                                     },
                  #sep=';'
                  #dtype=str
                 )


# In[26]:


data_9_head = op.load_workbook(r"Z:\Общая папка МИ по УД\Долговые карты субъектов\пункты на 01.09.2022\пункт 9\Расчет на 01.09.2022.xlsx", 
                                     data_only=True).active

head_g75 = data_9_head['A2'].value


# # ✅ Выгрузка пункт 10

# In[27]:


#Считываем за прошлый год файл пункт 10
data_10 = pd.read_excel(r"Z:\Общая папка МИ по УД\Долговые карты субъектов\пункты на 01.09.2022\пункт 10\п 10.xlsm", 
                               #sheet_name='1190',
                               header = 2,
                               names=['0.', 'region', '1.', '2.', '3.', '4.', '5.', '6.', '7.', '8.', '9.', 'reg'],
                               converters={'reg' : str,                              
                                     },
                  #sep=';'
                  #dtype=str
                 )

#Считываем за прошлый год файл пункт 10
data_10_KN = pd.read_excel(r"Z:\Общая папка МИ по УД\Долговые карты субъектов\пункты на 01.09.2022\пункт 10\п 10.xlsm", 
                               sheet_name='П10КН',
                               header = 4,
                               names=['0.', 'region', '1.', '2.', 'inn', 'kpp', '5.', '6.', '7.', '8.', '9.', '10.',
                                      '11.', '12.', '13.'],
                               converters={'inn' : str,  
                                           'kpp' : str,
                                     },
                  #sep=';'
                  #dtype=str
                 )


# In[28]:


data_10_KN.drop(columns=['0.', '1.'], 
                inplace=True
                )


# In[29]:


# в этом столбце было много нулей, даже там, где нет другой инфы и чтобы нули не вылетали в дашборд заменил на NaN
data_10['6.'] = data_10['6.'].replace({0 : np.nan})


# # ✅ Выгрузка Данные по руководству и телефоны

# In[30]:


nachalnik_data = pd.read_excel(r"Z:\Общая папка МИ по УД\Долговые карты субъектов\пункты на 01.04.2022\телефоны.xlsx", 
                               sheet_name='руководитель УФНС и МИ по КН',
                               header = 4,
                               names=['reg', 'region', '1.', '2.', '3.'],
                               converters={'reg' : str,                              
                                     },
                              )

zami_data = pd.read_excel(r"Z:\Общая папка МИ по УД\Долговые карты субъектов\пункты на 01.04.2022\телефоны.xlsx", 
                               sheet_name='заместители УФНС и МИ по КН',
                               header = 3,
                               names=['0.', 'reg', 'region', '1.', '2.', '3.', '4.', '5.', '6.', '7.', '8.', '9.',
                                     '10.', '11.'],
                               converters={'reg' : str,                              
                                     },
                         )
                          
otvetstvenie_data = pd.read_excel(r"Z:\Общая папка МИ по УД\Долговые карты субъектов\пункты на 01.04.2022\телефоны.xlsx", 
                               sheet_name='начальники отделов УЗ, РСБ',
                               header = 4,
                               names=['0.', 'reg', 'region', '1.', '2.', '3.', '4.', '5.', '6.', '7.', '8.', '9.',
                                     '10.', '11.', '12.', '13.', '14.', '15.', '16.'],
                               converters={'reg' : str,                              
                                     },
                                 )
                                  
DC_data = pd.read_excel(r"Z:\Общая папка МИ по УД\Долговые карты субъектов\пункты на 01.04.2022\телефоны.xlsx", 
                               sheet_name='начальники Долговых центров',
                               header = 2,
                              names=['reg', 'region', '1.', '2.', '3.', '4.', '5.', '6.', '7.', '8.', '9.',
                                     '10.', '11.', '12.', '13.'],
                               converters={'reg' : str,                              
                                     },
                       )


# # ✅ Уровни 

# In[31]:


#Считываем уровни
data_yroven = pd.read_excel(r"Z:\Общая папка МИ по УД\Долговые карты субъектов\Приложение 2.xlsx", 
                               #sheet_name='1190',
                               header = 6,
                               names=['reg', 'region', '1.', '2.', '3.'],
                               converters={'reg' : str,                              
                                     },
                  #sep=';'
                  #dtype=str
                 )


# # ✅ ВЫгрузка КАРТИНКИ

# In[32]:


path = r'Z:\Общая папка МИ по УД\Долговые карты субъектов\пункты на 01.09.2022\СКУАД'
all_files = glob.glob(path + "/*.png")


# In[33]:


path_2 = r'Z:\Общая папка МИ по УД\Долговые карты субъектов\долговые карты картинки'
all_files_2 = glob.glob(path_2 + "/*.png")


# # 🌐 проверка всех таблиц по столбцам, где есть названия регионов на лишние пробелы 

# In[34]:


data_new_5


# In[35]:


regions['region'] = regions['region'].map(str.strip)  
data_history['Unnamed: 1'] = data_history['Unnamed: 1'].map(str.strip) 
data_new['Unnamed: 1'] = data_new['Unnamed: 1'].map(str.strip) 
data_new_2['Unnamed: 2'] = data_new_2['Unnamed: 2'].map(str.strip) 
data_history_3['Unnamed: 0'] = data_history_3['Unnamed: 0'].map(str.strip) 
data_new_3['Unnamed: 0'] = data_new_3['Unnamed: 0'].map(str.strip) 
data_new_5['reg'] = data_new_5['reg'].map(str.strip) 
data_new_8['reg'] = data_new_8['reg'].map(str.strip) 
data_history_8['reg'] = data_history_8['reg'].map(str.strip)


# # 🌐 добавление дополн столбца с нумерацией региона

# In[36]:


# добавление пункту 2 столбцов с нумерацей регионов
data_new_2 = pd.merge(data_new_2,
                    regions, 
                    how='left', 
                    left_on='Unnamed: 2',
                    right_on = 'region')


data_new_2.drop(columns=['region'], inplace=True)

# проверка должно быть 85
data_new_2.info()


# In[37]:


# добавление пункту 3 столбцов с нумерацей регионов
data_history_3 = pd.merge(data_history_3,
                          regions, 
                          how='left', 
                          left_on='Unnamed: 0',
                         right_on = 'region')


data_history_3.drop(columns=['region'], inplace=True)
# проверка numb должно быть 95
#data_history_3.info()

data_new_3 = pd.merge(data_new_3,
                          regions, 
                          how='left', 
                          left_on='Unnamed: 0',
                         right_on = 'region')


data_new_3.drop(columns=['region'], inplace=True)
# проверка numb должно быть 95
#data_new_3.info()


# In[38]:


# добавление пункту 2 столбцов с нумерацей регионов
data_new_5 = pd.merge(data_new_5,
                    regions, 
                    how='left', 
                    left_on='reg',
                    right_on = 'region')


#data_new_5.drop(columns=['numb_y'], inplace=True)

# проверка должно быть 85
#data_new_5.info()


# In[39]:


# добавление пункту 3 столбцов с нумерацей регионов
data_history_8 = pd.merge(data_history_8,
                          regions, 
                          how='left', 
                          left_on='reg',
                         right_on = 'region')


data_history_8.drop(columns=['region'], inplace=True)
# проверка numb должно быть 95
#data_history_8.info()

data_new_8 = pd.merge(data_new_8,
                          regions, 
                          how='left', 
                          left_on='reg',
                         right_on = 'region')


data_new_8.drop(columns=['region'], inplace=True)
# проверка numb должно быть 95
#data_new_8.info()


# # 🌐 Обработка пункт 1

# In[40]:


# Так где указаано Росийская Федерация подставляем ункальное название РФ, чтобы дальше отбирать по этому названию
data_history.at[1, 'Unnamed: 0'] = 'RF'

# Так где указаано Росийская Федерация подставляем ункальное название РФ, чтобы дальше отбирать по этому названию
data_new.at[1, 'Unnamed: 0'] = 'RF'

# удаляем в первом столбце где NULL, чтобы в выгрузку не попадали не регионы, а пустые строки
data_history = data_history.dropna(subset = ['Unnamed: 0'])

# удаляем в первую сторчку
data_history = data_history.drop(labels = [0], axis = 0)

# удаляем в первом столбце где NULL, чтобы в выгрузку не попадали не регионы, а пустые строки
data_new = data_new.dropna(subset = ['Unnamed: 0'])

# удаляем в первую сторчку
data_new = data_new.drop(labels = [0], axis = 0)


data_history = data_history.sort_values(by='Unnamed: 0')
data_new = data_new.sort_values(by='Unnamed: 0')

# переименуем столбец для удобства
data_history.rename(columns = {'Unnamed: 0' : 'Unnamed'
                              }, inplace = True)

# переименуем столбец для удобства
data_new.rename(columns = {'Unnamed: 0' : 'Unnamed'
                              }, inplace = True)


# # 🌐 Обработка пункт 2

# In[41]:


# Так где указаано Росийская Федерация подставляем ункальное название РФ, чтобы дальше отбирать по этому названию
data_new_2.at[0, 'numb'] = 'RF'

# удаляем в первом столбце где NULL, чтобы в выгрузку не попадали не регионы, а пустые строки
#data_new = data_new.dropna(subset = ['Unnamed: 0'])

# удаляем в первую строчку
#data_history = data_history.drop(labels = [0], axis = 0)
data_new_2 = data_new_2.sort_values(by='numb')


data_new_2.drop(columns=['Unnamed: 3'], inplace=True)

display(data_new_2.head(2))


# # 🌐 Обработка пункт 3

# In[42]:


# Так где указаано Росийская Федерация подставляем ункальное название РФ, чтобы дальше отбирать по этому названию
data_history_3.at[1, 'numb'] = 'RF'

# Так где указаано Росийская Федерация подставляем ункальное название РФ, чтобы дальше отбирать по этому названию
data_new_3.at[1, 'numb'] = 'RF'

# удаляем в первом столбце где NULL, чтобы в выгрузку не попадали не регионы, а пустые строки
data_history_3 = data_history_3.dropna(subset = ['numb'])

# удаляем в первую сторчку
#data_history_3 = data_history_3.drop(labels = [0], axis = 0)

# удаляем в первом столбце где NULL, чтобы в выгрузку не попадали не регионы, а пустые строки
data_new_3 = data_new_3.dropna(subset = ['numb'])

# удаляем в первую сторчку
#data_new_3 = data_new_3.drop(labels = [0], axis = 0)


#data_history = data_history.sort_values(by='Unnamed: 0')
#data_new = data_new.sort_values(by='Unnamed: 0')

# переименуем столбец для удобства
data_history_3.rename(columns = {'Unnamed: 0' : 'Unnamed'
                              }, inplace = True)

# переименуем столбец для удобства
data_new_3.rename(columns = {'Unnamed: 0' : 'Unnamed'
                              }, inplace = True)

#удалим ненужные столбцы
data_history_3.drop(columns=[' Для РФ - 1-НМ (гр.2+гр.3)+4000 (гр.2);\nДля МИ по КН: 1-МР (гр.2+гр.3)+4000 (гр.2)\nДля регионов: (1-НМ (гр.2+гр.3)+4000 (гр. 2)) -(1-МР (гр.2+гр.3)+4000 (гр,2))',
                            ' Для РФ - 1-НМ (гр.2+гр.3)+4000 (гр.2)\nДля МИ по КН - 1-МР (гр.2+гр.3)\nДля регионов 1-НМ (гр.2+гр.3)+4000 (гр. 2)-1-МР (гр.2+гр.3))',
                            'Для РФ - 1-НМ (гр.2+гр.3)+4000 (гр.2)\nДля МИ по КН - 1-МР (гр.2+гр.3)\nДля регионов 1-НМ (гр.2+гр.3)+4000 (гр. 2)-1-МР (гр.2+гр.3))',
                            ], inplace=True)
data_new_3.drop(columns=[' Для РФ - 1-НМ (гр.2+гр.3)+4000 (гр.2);\nДля МИ по КН: 1-МР (гр.2+гр.3)+4000 (гр.2)\nДля регионов: (1-НМ (гр.2+гр.3)+4000 (гр. 2)) -(1-МР (гр.2+гр.3)+4000 (гр,2))',
                        ' Для РФ - 1-НМ (гр.2+гр.3)+4000 (гр.2)\nДля МИ по КН - 1-МР (гр.2+гр.3)\nДля регионов 1-НМ (гр.2+гр.3)+4000 (гр. 2)-1-МР (гр.2+гр.3))',
                        'Для РФ - 1-НМ (гр.2+гр.3)+4000 (гр.2)\nДля МИ по КН - 1-МР (гр.2+гр.3)\nДля регионов 1-НМ (гр.2+гр.3)+4000 (гр. 2)-1-МР (гр.2+гр.3))'], inplace=True)


# Добавляем строку с данными без крупнейших
punkt_3_bez_KN_old = {
        'Unnamed' : 'РОССИЙСКАЯ ФЕДЕРАЦИЯ (БЕЗ КН)',
        '№ 4-НМ (гр.1 стр.1001) ' : (data_history_3.iloc[0, 1] - data_history_3.iloc[1, 1] - data_history_3.iloc[2, 1] - data_history_3.iloc[3, 1] \
                                  - data_history_3.iloc[4, 1] - data_history_3.iloc[5, 1] - data_history_3.iloc[6, 1] - data_history_3.iloc[7, 1] \
                                  - data_history_3.iloc[8, 1] - data_history_3.iloc[9, 1] - data_history_3.iloc[10, 1]),
        '\nгр.3+гр.4-гр.5' : (data_history_3.iloc[0, 2] - data_history_3.iloc[1, 2] - data_history_3.iloc[2, 2] - data_history_3.iloc[3, 2] \
                            - data_history_3.iloc[4, 2] - data_history_3.iloc[5, 2] - data_history_3.iloc[6, 2] - data_history_3.iloc[7, 2] \
                            - data_history_3.iloc[8, 2] - data_history_3.iloc[9, 2] - data_history_3.iloc[10, 2]),
        'гр.2 / (гр.3 + гр. 4 - гр. 5) х 100\n' : (data_history_3.iloc[0, 1] - data_history_3.iloc[1, 1] - data_history_3.iloc[2, 1] - data_history_3.iloc[3, 1] \
                                  - data_history_3.iloc[4, 1] - data_history_3.iloc[5, 1] - data_history_3.iloc[6, 1] - data_history_3.iloc[7, 1] \
                                  - data_history_3.iloc[8, 1] - data_history_3.iloc[9, 1] - data_history_3.iloc[10, 1]) / 
                            (data_history_3.iloc[0, 2] - data_history_3.iloc[1, 2] - data_history_3.iloc[2, 2] - data_history_3.iloc[3, 2] \
                            - data_history_3.iloc[4, 2] - data_history_3.iloc[5, 2] - data_history_3.iloc[6, 2] - data_history_3.iloc[7, 2] \
                            - data_history_3.iloc[8, 2] - data_history_3.iloc[9, 2] - data_history_3.iloc[10, 2]),
        'numb' : 'BK'   # без крупнейших
}

punkt_3_bez_KN_new = {
        'Unnamed' : 'РОССИЙСКАЯ ФЕДЕРАЦИЯ (БЕЗ КН)',
        '№ 4-НМ (гр.1 стр.1001) ' : (data_new_3.iloc[0, 1] - data_new_3.iloc[1, 1] - data_new_3.iloc[2, 1] - data_new_3.iloc[3, 1] \
                                  - data_new_3.iloc[4, 1] - data_new_3.iloc[5, 1] - data_new_3.iloc[6, 1] - data_new_3.iloc[7, 1] \
                                  - data_new_3.iloc[8, 1] - data_new_3.iloc[9, 1] - data_new_3.iloc[10, 1]),
        '\nгр.3+гр.4-гр.5' : (data_new_3.iloc[0, 2] - data_new_3.iloc[1, 2] - data_new_3.iloc[2, 2] - data_new_3.iloc[3, 2] \
                            - data_new_3.iloc[4, 2] - data_new_3.iloc[5, 2] - data_new_3.iloc[6, 2] - data_new_3.iloc[7, 2] \
                            - data_new_3.iloc[8, 2] - data_new_3.iloc[9, 2] - data_new_3.iloc[10, 2]),
        'гр.2 / (гр.3 + гр. 4 - гр. 5) х 100\n' : (data_new_3.iloc[0, 1] - data_new_3.iloc[1, 1] - data_new_3.iloc[2, 1] - data_new_3.iloc[3, 1] \
                                  - data_new_3.iloc[4, 1] - data_new_3.iloc[5, 1] - data_new_3.iloc[6, 1] - data_new_3.iloc[7, 1] \
                                  - data_new_3.iloc[8, 1] - data_new_3.iloc[9, 1] - data_new_3.iloc[10, 1]) / 
                            (data_new_3.iloc[0, 2] - data_new_3.iloc[1, 2] - data_new_3.iloc[2, 2] - data_new_3.iloc[3, 2] \
                            - data_new_3.iloc[4, 2] - data_new_3.iloc[5, 2] - data_new_3.iloc[6, 2] - data_new_3.iloc[7, 2] \
                            - data_new_3.iloc[8, 2] - data_new_3.iloc[9, 2] - data_new_3.iloc[10, 2]),
        'numb' : 'BK'    # без крупнейших
        }


data_history_3 = data_history_3.append(punkt_3_bez_KN_old, ignore_index=True).sort_values(by='numb')
data_new_3 = data_new_3.append(punkt_3_bez_KN_new, ignore_index=True).sort_values(by='numb')


# In[43]:


data_3 = pd.merge(data_history_3,
                   data_new_3, 
                   how='left', 
                   on = 'numb')
data_3.drop(columns=['Unnamed_y'], inplace=True)
data_3 = data_3[['numb', 'Unnamed_x', '№ 4-НМ (гр.1 стр.1001) _x', '\nгр.3+гр.4-гр.5_x', 
                'гр.2 / (гр.3 + гр. 4 - гр. 5) х 100\n_x', '№ 4-НМ (гр.1 стр.1001) _y',
                '\nгр.3+гр.4-гр.5_y', 'гр.2 / (гр.3 + гр. 4 - гр. 5) х 100\n_y']]


# # 🌐 Обработка пункт 4

# In[44]:


data_new_4.tail(4)


# In[45]:


def to_int(data_f, colum):
    data_f[colum] = data_f[colum].astype('int')
    #data_f[colum] = data_f[colum].replace(',', '.', regex=True).astype('int')

#to_int(data_new_4, 'col_3')
#to_int(data_new_4, 'col_4')
#to_int(data_new_4, 'col_5')
#to_int(data_new_4, 'col_6')
#to_int(data_new_4, 'col_7')
#to_int(data_new_4, 'col_8')
#to_int(data_new_4, 'col_9')
#to_int(data_new_4, 'col_10')
#to_int(data_new_4, 'col_11')
#to_int(data_new_4, 'col_12')
#to_int(data_new_4, 'col_13')


data_new_4['col_4'] = data_new_4['col_4'] #/ 100
data_new_4['col_5'] = data_new_4['col_5'] #/ 100
data_new_4['col_6'] = data_new_4['col_6'] #/ 100


# In[46]:


data_new_4.at[86, 'numb'] = 'RF'


# # 🌐 Обработка пункт 5

# In[47]:


# удаляем в первом столбце где NULL, чтобы в выгрузку не попадали не регионы, а пустые строки
for y in range(len(data_5_left)-1):
    data_5_left[y] = data_5_left[y].dropna(subset = ['number'])


# In[48]:


#перведем в int некоторые столбцы
def to_int_2(data_f, colum):
    data_f[colum] = data_f[colum].astype('int')
    #data_f[colum] = data_f[colum].replace(',', '.', regex=True).astype('int')

for a in range(len(data_5_left)-1):
    to_int_2(data_5_left[a], 'number')


# In[49]:


# категоризировали зонам. позиция с 1 по 15 - хорошо, с 16 по 42 - нейтральная зона, 
#     с 43 по 69 - зона риска, 70 по 84 - плохо
def zona_categ(row):
    zona = row['number']
    
    if 0 < zona <= 15:
        return 'хорошо'
    elif 15 < zona <= 42:
        return 'нейтральная зона'
    elif 15 < zona <= 69:
        return 'зона риска'
    elif 69 < zona <= 84:
        return 'плохо'

    
#data_new_5['zona'] = data_new_5.apply(zona_categ, axis=1)
for h in range(len(data_5_left)-1):
    data_5_left[h]['zona'] = data_5_left[h].apply(zona_categ, axis=1) 


# In[50]:


data_new_5 = data_new_5.dropna(subset = ['numb_x'])


# In[51]:


#перведем в int некоторые столбцы
#def to_int_2(data_f, colum):
    #data_f[colum] = data_f[colum].astype('int')
    #data_f[colum] = data_f[colum].replace(',', '.', regex=True).astype('int')

#to_int_2(data_new_5, 'numb_x')
#to_int_2(data_new_5, '5.')
#to_int_2(data_new_5, '15.')
#to_int_2(data_new_5, '19.')
#to_int_2(data_new_5, '29.')
#to_int_2(data_new_5, '39.')
#to_int_2(data_new_5, '43.')
#to_int_2(data_new_5, '47.')
#to_int_2(data_new_5, '51.')
#to_int_2(data_new_5, '62.')
#to_int_2(data_new_5, '66.')


# In[52]:


data_new_5['23.'] = data_new_5['23.'] * 100


# In[53]:


# округлим некоторые столбцы до 1 знака
def to_round_1(data_f, colum):
    data_f[colum] = data_f[colum].round(1)
    #data_f[colum] = data_f[colum].replace(',', '.', regex=True).astype('int')

to_round_1(data_new_5, '8.')
to_round_1(data_new_5, '14.')
to_round_1(data_new_5, '33.')
to_round_1(data_new_5, '39.')
to_round_1(data_new_5, '47.')
to_round_1(data_new_5, '57.')
to_round_1(data_new_5, '63.')
to_round_1(data_new_5, '72.')
to_round_1(data_new_5, '23.')
#to_round_1(data_new_5, '65.')


# In[54]:


general_5_left = pd.merge(data_5_left[0],
                   data_5_left[1], 
                   how='inner', 
                   on = 'regions'
                         )


# In[55]:


for t in range(2, len(data_5_left)):
    general_5_left = general_5_left.merge(data_5_left[t], 
                   how='left', 
                   on = 'regions'
                                         )
    


# In[56]:


# добавление пункту 2 столбцов с нумерацей регионов
general_5_left = pd.merge(general_5_left,
                    regions, 
                    how='left', 
                    left_on='regions',
                    right_on = 'region')


# In[57]:


general_5_left.drop(columns=['region', 'okrug'], inplace=True)


# In[58]:


general_5_left = general_5_left.sort_values(by='numb')


# # 🌐 Обработка пункт 6

# In[59]:


# Так где указаано Росийская Федерация подставляем ункальное название РФ, чтобы дальше отбирать по этому названию
data_history_6.at[1, 'reg'] = 'RF'

# Так где указаано Росийская Федерация подставляем ункальное название РФ, чтобы дальше отбирать по этому названию
data_new_6.at[1, 'reg'] = 'RF'

# удаляем в первом столбце где NULL, чтобы в выгрузку не попадали не регионы, а пустые строки
data_history_6 = data_history_6.dropna(subset = ['reg'])


# удаляем в первом столбце где NULL, чтобы в выгрузку не попадали не регионы, а пустые строки
data_new_6 = data_new_6.dropna(subset = ['reg'])


# In[60]:


data_history_6 = data_history_6[['reg', 'region', '1.', '2.', '3.', '4.', '5.', '6.', '7.', '8.', '9.',
       '10.', '11.', '12.', '13.', '14.', '15.', '19.',
       '20.', '21.', '22.', '23.', '24.']]


# In[61]:


data_new_6 = data_new_6[['reg', 'region', '1.', '2.', '3.', '4.', '5.', '6.', '7.', '8.', '9.',
       '10.', '11.', '12.', '13.', '14.', '15.', '22.', '23.', '24.']]


# In[62]:


data_history_6.head(3)


# In[63]:


data_new_6.head(3)


# In[64]:


# абсолютные значения привести к процентным
def procent_100(dt, col):
    dt[col] = dt[col] * 100
    
procent_100(data_history_6, '3.') 
procent_100(data_history_6, '6.')  
procent_100(data_history_6, '9.')   
procent_100(data_history_6, '12.') 
procent_100(data_history_6, '15.')  
procent_100(data_history_6, '21.')   
#procent_100(data_new_6, '3.') 
#procent_100(data_new_6, '6.')  
#procent_100(data_new_6, '9.')   
#procent_100(data_new_6, '12.') 
#procent_100(data_new_6, '15.')  
#procent_100(data_new_6, '21.')  


# # 🌐 Обработка пункт 7

# In[65]:


data_general_7 = {}


# In[66]:


# Должен быть 0
len(data_7) - len(list_region_7)


# In[67]:


for i in range(len(list_region_7)):
    data_general_7[list_region_7[i]] = data_7[i] 


# In[68]:


#перведем в int некоторые столбцы
def to_int_7(data_f, colum):
    data_f[colum] = data_f[colum].astype('int')

    
for a in (data_general_7.keys()):
    #print(str(a))
    to_int_7(data_general_7[a], 'all')
    to_int_7(data_general_7[a], 'UL')
    to_int_7(data_general_7[a], 'FL')
    to_int_7(data_general_7[a], 'IP')


# # 🌐 Обработка пункт 8

# In[69]:


# удаляем в первом столбце где NULL, чтобы в выгрузку не попадали не регионы, а пустые строки
data_history_8 = data_history_8.dropna(subset = ['numb'])


# удаляем в первом столбце где NULL, чтобы в выгрузку не попадали не регионы, а пустые строки
data_new_8 = data_new_8.dropna(subset = ['numb'])


# # 🌐 Обработка пункт 9

# In[70]:


data_9.at[0, 'region'] = 'Российская Федерация'

# удаляем в первом столбце где NULL, чтобы в выгрузку не попадали не регионы, а пустые строки
data_9 = data_9.dropna(subset = ['reg'])

data_9.head(3)


# # 🌐 Обработка пункт 10

# In[71]:


# добавление пункту 10 по КН столбцов с нумерацей кода НО
data_10_KN_general = pd.merge(data_10_KN,
                    regions, 
                    how='left', 
                    left_on='region',
                    right_on = 'region')


# In[72]:


data_10_KN_general = data_10_KN_general.dropna(subset = ['inn'])


# In[73]:


def to_int_10(data_f, colum):
    data_f[colum] = data_f[colum].astype('int')
    #data_f[colum] = data_f[colum].replace(',', '.', regex=True).astype('int')

to_int_10(data_10_KN_general, '11.')


# In[74]:


# перевод datatime в date
def datet_10(dt, col):
    dt[col] = pd.to_datetime(dt[col],
                             dayfirst=True,
                             #format='%d%m%Y'
                            ).dt.strftime('%m/%Y')


datet_10(data_10_KN_general, '8.')


# In[75]:


data_10_KN_general['7.'] = data_10_KN_general['7.'].round(3)


# In[76]:


data_10_KN_general


# # 🌐 Обработка УРОВНИ

# In[77]:


# добавил столбец с укроченным названием, чтобы потом по нему заполнять
data_yroven['4.'] = data_yroven['3.']

for i in range (len(data_yroven)):    
    data_yroven['4.'][i] = data_yroven.iloc[i, 4][:4]


# # 🌐 Обработка КАРТИНКИ

# # 💯Сбор общего файла по регионам

# In[83]:


data_22 = ['01', '07', '03']


# In[84]:


otvetstvenie_data.head(14)


# # 🅰 Разбивка УФНС

# In[99]:


#j, q, w, e = 0, 0, 1, 2

for i in data_22:
    #j = 0
    # пункт 1️⃣
    P_1_history = data_history.query('Unnamed == "RF" | Unnamed == @i') # пункт 1 история
    P_1_new= data_new.query('Unnamed == "RF" | Unnamed == @i')          # пункт 1 новое
    P_1_history = P_1_history.drop(['Unnamed'], axis = 1)
    P_1_new = P_1_new.drop(['Unnamed'], axis = 1)
    P_1_new.rename(columns = {'Unnamed: 1' : ''}, inplace = True)
    P_1_history = P_1_history.sort_index(ascending=True)
    P_1_new = P_1_new.sort_index(ascending=True)
    
    # пункт 2️⃣
    P_2_new = data_new_2.query('numb == "RF" | numb == @i')
    P_2_new = P_2_new.sort_index(ascending=True)
    
    # пункт 3️⃣
    P_3 = data_3.query('numb == "RF" | numb == "BK" | numb == @i') # пункт 3
    P_3 = P_3.sort_index(ascending=False)   
    
    # пункт 4️⃣
    P_4 = data_new_4.query('numb == "RF" | numb == @i') # пункт 4   
    P_4 = P_4.sort_index(ascending=True)
    
    # пункт 5️⃣
    P_5 = data_new_5.query('numb_y == @i') # пункт 5  
    P_5 = P_5.sort_index(ascending=True)
    
    p_5_left = general_5_left.query('numb == @i')
    
    # пункт 6️⃣
    P_6_history = data_history_6.query('reg == "RF" | reg == @i') # пункт 6 история
    P_6_new= data_new_6.query('reg == "RF" | reg == @i')          # пункт 6 новое
    P_6_history = P_6_history.sort_index(ascending=True)
    P_6_new = P_6_new.sort_index(ascending=True)
    
    # пункт 7️⃣
    P_7 = data_general_7[i]     
    
    # пункт 8️⃣
    P_8_history = data_history_8.query('numb == @i')     # пункт 8 история
    P_8_new= data_new_8.query('numb == @i')             # пункт 8 новое
    
    # пункт 9️⃣
    P_9 = data_9.query('reg == "Российская Федерация" | reg == @i') # пункт 9
    P_9 = P_9.sort_index(ascending=True)  
    
    # пункт 🔟
    P_10 = data_10.query('reg == @i')
    
# НАЧАЛЬНИК
    P_nachalnik = nachalnik_data.query('reg == @i') 
    # ЗАМЫ
    P_zami = zami_data.query('reg == @i')
    # ОТветственные
    P_otvetstvenie = otvetstvenie_data.query('reg == @i') 
    # Долговые центры
    P_DC = DC_data.query('reg == @i')

# уровни   
    P_yrovni = data_yroven.query('reg== @i') # пункт уровни

    # КАРТИНКИ
    #img = Image.open(all_files[int(i) - 1]) 
    #img.show()
    
    
    shutil.copy(r"Z:\8-Отдел анализа данных\001 - Сотрудники\05 - Чурахин А.А\Долговая карта Python\Шаблон новый 01.11.xlsx", 
                i+'.xlsx')
    wb = op.load_workbook(i+'.xlsx')
    #ws1 = wb.create_sheet('Help', 1)
    ws = wb.active
    #ws1 = wb.active
    
###  ШАПКА ОБЩАЯ НАЗВАНИЕ РЕГИОНА И ОКРУГ
    ws['A4'] = P_2_new.iloc[1, 2]
    ws['A5'] = P_2_new.iloc[1, 12]

#  заносим НАЧАЛЬНИК
    ws['H2'] = P_nachalnik.iloc[0, 2]             # ФИО
    ws['I4'] = P_nachalnik.iloc[0, 3]             # телефон
    
#  заносим ЗАМЫ
    for row, l in zip(['K1', 'K2', 'N1', 'M2'], [13, 4, 6, 5]):
        ws[row] = P_zami.iloc[0, l]  
#  заносим ЗАМЫ  
    for row, l in zip(['O1', 'O2', 'R1', 'Q2'], [13, 4, 6, 5]):
        ws[row] = P_zami.iloc[1, l]
# заносим ЗАМЫ    
    for row, l in zip(['S1', 'S2', 'V1', 'U2'], [13, 4, 6, 5]):
        ws[row] = P_zami.iloc[2, l]
        
    #display(P_otvetstvenie)    
#  заносим Ответственные отделы
    #for row, l in zip(['R68', 'S68', 'T68', 'R69', 'S69', 'T69'], [6, 5, 4, 17, 16, 15]):
    try:
        ws['K4'] = P_otvetstvenie.iloc[0, 4] 
        ws['K5'] = P_otvetstvenie.iloc[0, 6] + ' - ' + P_otvetstvenie.iloc[0, 5]
        ws['K7'] = P_otvetstvenie.iloc[0, 17] + '\n' + P_otvetstvenie.iloc[0, 16]
    except:
        pass
    try:
        ws['N4'] = P_otvetstvenie.iloc[1, 4] 
        ws['N5'] = P_otvetstvenie.iloc[1, 6] + ' - ' + P_otvetstvenie.iloc[1, 5]
        ws['N7'] = P_otvetstvenie.iloc[1, 17] + '\n' + P_otvetstvenie.iloc[1, 16]       
    except:
        pass
    try:
        ws['Q4'] = P_otvetstvenie.iloc[2, 4] 
        ws['Q5'] = P_otvetstvenie.iloc[2, 6] + ' - ' + P_otvetstvenie.iloc[2, 5]
        ws['Q7'] = P_otvetstvenie.iloc[2, 17]+ '\n' +P_otvetstvenie.iloc[2, 16]
    except:
        pass
    try:
        ws['T4'] = P_otvetstvenie.iloc[3, 4]
        ws['T5'] = P_otvetstvenie.iloc[3, 6] + ' - ' + P_otvetstvenie.iloc[3, 5]
        ws['T7'] = P_otvetstvenie.iloc[3, 17] + '\n' + P_otvetstvenie.iloc[3, 16]
    except:
        pass
    
#  заносим Ответственные отделы 
    #for row, l in zip(['R72', 'S72', 'T72', 'R73', 'S73', 'T73'], [6, 5, 4, 17, 16, 15]):
        #ws[row] = P_otvetstvenie.iloc[1, l]
# заносим Ответственные отделы     
    #for row, l in zip(['R75', 'S75', 'T75', 'R76', 'S76', 'T76'], [6, 5, 4, 17, 16, 15]):
        #ws[row] = P_otvetstvenie.iloc[2, l]  
# заносим Ответственные отделы     
    #for row, l in zip(['R78', 'S78', 'T78', 'R79', 'S79', 'T79'], [6, 5, 4, 17, 16, 15]):
        #ws[row] = P_otvetstvenie.iloc[3, l]         

#  заносим Долговые центры
    #ws['R21'] = 'Долговой центр создан'
    try:
        for row, l in zip(['R22', 'S22', 'R23', 'S23', 'T23', 'S24'], [3, 2, 13, 5, 14, 12]):
            ws[row] = P_DC.iloc[0, l]
            ws['R21'] = 'Долговой центр создан'
    except: pass
    
#  заносим уровни
    #display(P_yrovni.iloc[0, 5])
    if P_yrovni.iloc[0, 5] == 'двух':
        ws['L21'] = 'Двухуровневая система'
    elif P_yrovni.iloc[0, 5] == 'трех':
        ws['L21'] = 'Трехуровневая система'
    
#  заносим шапку пункт 1️⃣
    ws['C11'] = head_c11
    ws['G11'] = head_g11
    ws['K11'] = head_k11 + ' на ' + head_c11[2:] + ' к сведению на ' + head_g11[2:]
    ws['C16'] = head_c16
    ws['G16'] = head_g16
    ws['K16'] = head_k16 + ' на ' + head_c16[2:] + ' к сведению на ' + head_g16[2:]  
    
  #  заносим данные пункт 1 строчка 1
    for row, l in zip(['A14', 'C14', 'D14', 'E14', 'F14', 'G14', 'H14', 'I14', 'J14', 'K14', 'L14', 'M14', 'N14', 'O14', 'P14', 'Q14', 'R14' ], range(0, 17)):
        ws[row] = P_1_history.iloc[0, l]  
  #  заносим данные пункт 1 строчка 2   
    for row, l in zip(['A15', 'C15', 'D15', 'E15', 'F15', 'G15', 'H15', 'I15', 'J15', 'K15', 'L15', 'M15', 'N15', 'O15', 'P15', 'Q15', 'R15' ], range(0, 17)):
        ws[row] = P_1_history.iloc[1, l]
# закрашивание        
    for row, l in zip(['L15', 'N15', 'P15', 'R15'], range(10, 17, 2)):
        if P_1_history.iloc[0, l] > P_1_history.iloc[1, l]:
            ws[row].fill = PatternFill(fill_type='solid', start_color = '54bd55')
        else: ws[row].fill = PatternFill(fill_type='solid', start_color = 'ed2213') 
        
  # заносим данные пункт 1 строчка 3      
    for row, l in zip(['A17', 'C17', 'D17', 'E17', 'F17', 'G17', 'H17', 'I17', 'J17', 'K17', 'L17', 'M17', 'N17', 'O17', 'P17', 'Q17', 'R17' ], range(0, 17)):
        ws[row] = P_1_new.iloc[0, l] 
  # заносим данные пункт 1 строчка 4  
    for row, l in zip(['A18', 'C18', 'D18', 'E18', 'F18', 'G18', 'H18', 'I18', 'J18', 'K18', 'L18', 'M18', 'N18', 'O18', 'P18', 'Q18', 'R18' ], range(0, 17)):
        ws[row] = P_1_new.iloc[1, l]
# закрашивание        
    for row, l in zip(['L18', 'N18', 'P18', 'R18' ], range(10, 17, 2)):
        if P_1_new.iloc[0, l] > P_1_new.iloc[1, l]:
            ws[row].fill = PatternFill(fill_type='solid', start_color = '54bd55')
        else: ws[row].fill = PatternFill(fill_type='solid', start_color = 'ed2213')    
        
      

  #  заносим шапку пункт 2️⃣ 
    ws['C21'] = 'на ' + head_c21
    ws['E21'] = 'на ' + head_e21   
    ws['G21'] = 'ДИНАМИКА ' + 'на ' + head_e21 + ' к ' + head_c21 
    
    #  заносим данные пункт 2 строчка 1
    for row, l in zip(['A24', 'C24', 'D24', 'E24', 'F24', 'G24', 'H24', 'I24', 'J24' ], range(2, 11)):
        ws[row] = P_2_new.iloc[0, l]
    #  заносим данные пункт 2 строчка 2    
    for row, l in zip(['A25', 'C25', 'D25', 'E25', 'F25', 'G25', 'H25', 'I25', 'J25' ], range(2, 11)):
        ws[row] = P_2_new.iloc[1, l]
        
# закрашивание         
    for row, l in zip(['H25', 'J25' ], [8, 10]):
        if P_2_new.iloc[0, l] > P_2_new.iloc[1, l]:
            ws[row].fill = PatternFill(fill_type='solid', start_color = '54bd55')
        else: ws[row].fill = PatternFill(fill_type='solid', start_color = 'ed2213')      
       
  
   #  заносим шапку пункт 3️⃣
    ws['C30'] = head_c30
    ws['F30'] = head_f30    
   #  заносим данные пункт 3 строчка 1
    for row, l in zip(['A32', 'C32', 'D32', 'E32', 'F32', 'G32', 'H32'], range(1, 8)):
        ws[row] = P_3.iloc[0, l]     
   #  заносим данные пункт 3 строчка 2
    for row, l in zip(['A33', 'C33', 'D33', 'E33', 'F33', 'G33', 'H33'], range(1, 8)):
        ws[row] = P_3.iloc[1, l]    
  #  заносим данные пункт 3 строчка 3      
    for row, l in zip(['A34', 'C34', 'D34', 'E34', 'F34', 'G34', 'H34'], range(1, 8)):
        ws[row] = P_3.iloc[2, l] 
        
# закрашивание         
    for row, l in zip(['E34', 'H34' ], [4, 7]):
        if P_3.iloc[0, l] > P_3.iloc[2, l]:
            ws[row].fill = PatternFill(fill_type='solid', start_color = '54bd55')
        else: ws[row].fill = PatternFill(fill_type='solid', start_color = 'ed2213')  
            
    
 
    #  заносим шапку пункт 4️⃣    В САМОЙ ТАБЛИЦЕ ШАПКИ НЕТ, ВЗЯЛ ЧУЖУЮ ДАТУ
    ws['K28'] = '4. Индикативный план на ' + head_f30    
    #ws1['A1'] = '4. Индикативный план на'
    #  заносим данные пункт 4 строчка 1
    for row, l in zip(['P29', 'P30', 'P31', 'P32', 'P33', 'P34', 'V29', 'V30', 'V31', 'V32', 'V33'], range(3, 13)):
        ws[row] = P_4.iloc[0, l]
    

# закрашивание         
    for row, l in zip(['V29', 'V30', 'V32'], [8, 9, 11]):
        if P_4.iloc[1, l] > P_4.iloc[0, l]:
            ws[row].fill = PatternFill(fill_type='solid', start_color = '54bd55')
        else: ws[row].fill = PatternFill(fill_type='solid', start_color = 'ed2213') 
            
            
      
   
    # заносим шапку пункт 5️⃣    
    ws['I38'] = '5. Рейтинг на ' + head_I38 + 'в разрезе показателей' 
    # вытаскиваваем значения которые будут идти вместе с текстом в правой части таблицы 5     
    ws['Q39'] = 'Эффективность работы с долгом по контрольной работе ' + str(P_5.iloc[0, 2])    
    ws['Q40'] = 'Динамика снижения (роста) совокупной задолженности к аналогичному периоду ' + str(P_5.iloc[0, 4])
    ws['Q41'] = 'Качество работы технологических процессов по управлению долгом, Кутп = ' + str(P_5.iloc[0, 6])+'%'
    ws['Q42'] = 'Динамика снижения (роста) DTI ' + str(P_5.iloc[0, 9])
    ws['Q43'] = 'Эффективность применения мер взыскания задолженности в отношении ЮЛ и ИП ' + str(P_5.iloc[0, 11])
    ws['Q44'] = 'Динамика применения ареста имущества налогоплательщиказадолженности ' + str(P_5.iloc[0, 13])
    ws['Q45'] = 'Эффективность работы с задолженностью ФЛ ' + str(P_5.iloc[0, 16])
    ws['Q46'] = 'Работа с дебиторской задолженностью должника ' + str(P_5.iloc[0, 18])
    ws['Q47'] = 'Эффективность работы с невыясненными платежами ' + str(P_5.iloc[0, 20])
    ws['Q48'] = 'Эф-ть работы по привлечению налогопла-в к представлению согласий на информирование о долге  '
    ws['Q49'] = 'Урег-е задол-и по рег-м и местным налогам, а также по НДФЛ путем предоставления отсрочки   ' 
    ws['Q50'] = 'Средний срок принятия решения о возврате излишне уплаченных налогов, страх. взносов   ' 
    # левая часть таблицы
    #ws['F39'] = P_5.iloc[0, 25]
    #ws['G39'] = P_5.iloc[0, 1]
    
##🚩 КАЖДЫЙ МЕСЯЦ НАДО ДОБАВЛЯТЬ ЗАПИСИ !!!!!!🚨!!!!!!!!!!!!!!!!!!!!!!!!!
    for row, l in zip(['C39', 'B39', 'C40', 'B40', 'C41', 'B41', 'C42', 'B42', 'C43', 'B43', 'C44', 'B44', 
                       'C45', 'B45', 'C46', 'B46', 'C47', 'B47', 'C48', 'B48', 'C49', 'B49', 'C50', 'B50', 
                       'G39', 'F39', 'G40', 'F40', 'G42', 'F42', 'G43', 'F43', 'G44', 'F44', 'G45', 'F45',
                       'G46', 'F46'], range(1, 39)):
        ws[row] = p_5_left.iloc[0, l]
        
    
##🚩 КАЖДЫЙ МЕСЯЦ НАДО ДОБАВЛЯТЬ ЗАПИСИ !!!!!!🚨!!!!!!!!!!!!!!!!!!!!!!!!!
#  левая часть     
    for row, l in zip(['C39', 'C40', 'C41', 'C42', 'C43', 'C44', 'C45', 'C46', 'C47', 'C48', 'C49', 'C50', 
    'G39', 'G40', 'G42', 'G43', 'G44', 'G45', 'G46'], range(1, 39, 2)):  
        if p_5_left.iloc[0, l] < 16:
            ws[row].fill = PatternFill(fill_type='solid', start_color = '5bc271')
        elif 16 <= p_5_left.iloc[0, l] < 43:
            ws[row].fill = PatternFill(fill_type='solid', start_color = 'e5a23c')
        elif 43 <= p_5_left.iloc[0, l] < 70:
            ws[row].fill = PatternFill(fill_type='solid', start_color = 'c24611')
        elif p_5_left.iloc[0, l] >= 70:
            ws[row].fill = PatternFill(fill_type='solid', start_color = 'e82309') 
            
            
##🚩 КАЖДЫЙ МЕСЯЦ НАДО ДОБАВЛЯТЬ ЗАПИСИ !!!!!!🚨!!!!!!!!!!!!!!!!!!!!!!!!!
    #  заносим данные пункт 5 строчка 1
    for row, l in zip(['P39', 'P40', 'P41', 'P42', 'P43', 'P44', 'P45', 'P46', 'P47', 'P48'], [3, 5, 7, 10, 12, 14, 17, 19, 21]):
        ws[row] = P_5.iloc[0, l]
        
        
#  правая часть      
    for row, l in zip(['P39', 'P40', 'P41', 'P42', 'P43', 'P44', 'P45', 'P46', 'P47', 'P48'], [3, 5, 7, 10, 12, 14, 17, 19, 21]):  
        if P_5.iloc[0, l] < 16:
            ws[row].fill = PatternFill(fill_type='solid', start_color = '5bc271')
        elif 16 <= P_5.iloc[0, l] < 43:
            ws[row].fill = PatternFill(fill_type='solid', start_color = 'e5a23c')
        elif 43 <= P_5.iloc[0, l] < 70:
            ws[row].fill = PatternFill(fill_type='solid', start_color = 'c24611')
        elif P_5.iloc[0, l] >= 70:
            ws[row].fill = PatternFill(fill_type='solid', start_color = 'e82309')       
    
    
    #  заносим шапку пункт 6️⃣    
    ws['A55'] = 'на 01.01.2022 г.' #head_a55  в шапке указан январь, а не номер месяца 
    ws['A61'] = head_a61 
 #  заносим данные пункт 6 строчка 1
    for row, l in zip(['A59', 'C59', 'D59', 'E59', 'F59', 'G59', 'H59', 'I59', 'J59', 'K59', 'L59', 'M59', 'N59', 'O59', 'P59', 'Q59', 'R59', 'S59', 'T59' ], range(1, 25)):
        ws[row] = P_6_history.iloc[0, l]  
  #  заносим данные пункт 6 строчка 2   
    for row, l in zip(['A60', 'C60', 'D60', 'E60', 'F60', 'G60', 'H60', 'I60', 'J60', 'K60', 'L60', 'M60', 'N60', 'O60', 'P60', 'Q60', 'R60', 'S60', 'T60'], range(1, 25)):
        ws[row] = P_6_history.iloc[1, l]
  # заносим данные пункт 6 строчка 3      
    for row, l in zip(['A62', 'C62', 'D62', 'E62', 'F62', 'G62', 'H62', 'I62', 'J62', 'K62', 'L62', 'M62', 'N62', 'O62', 'P62', 'Q62', 'R62', 'S62', 'T62'], range(1, 25)):
        ws[row] = P_6_new.iloc[0, l] 
  # заносим данные пункт 6 строчка 4 
    #display(P_6_new)
    for row, l in zip(['A63', 'C63', 'D63', 'E63', 'F63', 'G63', 'H63', 'I63', 'J63', 'K63', 'L63', 'M63', 'N63', 'O63', 'P63', 'Q63', 'R63', 'S63', 'T63'], range(1, 25)):
        ws[row] = P_6_new.iloc[1, l] 
        
# закрашивание         
    for row, l in zip(['E60', 'H60', 'K60', 'N60', 'Q60', 'T60'], [4, 7, 10, 13, 16, 19]):
        if P_6_history.iloc[1, l] > P_6_history.iloc[0, l]:
            ws[row].fill = PatternFill(fill_type='solid', start_color = '54bd55')
        else: ws[row].fill = PatternFill(fill_type='solid', start_color = 'ed2213')   
# закрашивание         
    for row, l in zip(['E63', 'H63', 'K63', 'N63', 'Q63', 'T63'], [4, 7, 10, 13, 16, 19]):
        if P_6_new.iloc[1, l] > P_6_new.iloc[0, l]:
            ws[row].fill = PatternFill(fill_type='solid', start_color = '54bd55')
        else: ws[row].fill = PatternFill(fill_type='solid', start_color = 'ed2213') 
            
     #  заносим шапку пункт 7️⃣  
    ws['A66'] = '7. Количество налогоплательщиков, имеющих задолженность по налогам, сборам,     страховым взносам, пеням и налоговым санкциям на ' + head_a61 
    #  заносим данные пункт 7 строчка 1
    for row, l in zip(['C71', 'D71'], [2, 3]):
        ws[row] = P_7.iloc[0, l] 
    ws['E71'] = P_7.iloc[0, 4] + P_7.iloc[0, 5]  
  #  заносим данные пункт 7 строчка 2   
    for row, l in zip(['C72', 'D72'], [2, 3]):
        ws[row] = P_7.iloc[1, l]
    ws['E72'] = P_7.iloc[1, 4] + P_7.iloc[1, 5]     
  # заносим данные пункт 7 строчка 3      
    for row, l in zip(['C73', 'D73'], [2, 3]):
        ws[row] = P_7.iloc[2, l] 
    ws['E73'] = P_7.iloc[2, 4] + P_7.iloc[2, 5]     
  # заносим данные пункт 7 строчка 4     
    for row, l in zip(['C74', 'D74'], [2, 3]):
        ws[row] = P_7.iloc[3, l] 
    ws['E74'] = P_7.iloc[3, 4] + P_7.iloc[3, 5]     
# заносим данные пункт 7 строчка 5      
    for row, l in zip(['C75', 'D75'], [2, 3]):
        ws[row] = P_7.iloc[4, l]
    ws['E75'] = P_7.iloc[4, 4] + P_7.iloc[4, 5]     
  # заносим данные пункт 7 строчка 6     
    for row, l in zip(['C76', 'D76'], [2, 3]):
        ws[row] = P_7.iloc[5, l]  
    ws['E76'] = P_7.iloc[5, 4] + P_7.iloc[5, 5]     
    
      
    
  #  заносим шапку пункт 8️⃣   
    ws['G69'] = 'на ' + head_g69 
    ws['I69'] = 'на ' + head_i69  
    ws['K69'] = 'Динамика на ' + head_g69 + ' к ' + head_i69 + ', %'
    ws['M69'] = 'на ' + head_g69 
    ws['N69'] = 'на ' + head_i69 
    ws['O69'] = 'на ' + head_g69 + ' к ' + head_i69 + ', %'
  #  заносим данные пункт 8 
    for row, l in zip(['I71', 'J71'], [1, 3]):
        ws[row] = P_8_history.iloc[0, l]  
  #  заносим данные пункт 8  
    for row, l in zip(['G71', 'H71'], [1, 3]):
        ws[row] = P_8_new.iloc[0, l]      
    for row, l in zip(['K71', 'L71'], [1, 3]):    
        ws[row] = ((P_8_new.iloc[0, l] / P_8_history.iloc[0, l]) * 100) - 100
# закрашивание         
    for row, l in zip(['K71', 'L71'], [1, 3]):
        if (((P_8_new.iloc[0, l] / P_8_history.iloc[0, l]) * 100) - 100) > 0:
            ws[row].fill = PatternFill(fill_type='solid', start_color = '54bd55')
        elif (((P_8_new.iloc[0, l] / P_8_history.iloc[0, l]) * 100) - 100) < 0:
            ws[row].fill = PatternFill(fill_type='solid', start_color = 'ed2213')   
                
       
    #  заносим шапку пункт 9️⃣   
    ws['G75'] = head_g75     
    #  заносим данные пункт 9
    for row, l in zip(['G78', 'M78', 'N78', 'O78', 'P78'], [1, 5, 2, 3, 4]):
        ws[row] = P_9.iloc[0, l]
    for row, l in zip(['G79', 'I79', 'M79', 'N79', 'O79', 'P79'], [0, 1, 5, 2, 3, 4]):
        ws[row] = P_9.iloc[1, l]
        
    for row, l in zip([ 'M79', 'N79', 'O79', 'P79'], [5, 2, 3, 4]):
        if P_9.iloc[0, l] <  P_9.iloc[1, l]:
            ws[row].fill = PatternFill(fill_type='solid', start_color = '54bd55')
        else: ws[row].fill = PatternFill(fill_type='solid', start_color = 'ed2213')   
            
            
    
#  заносим данные пункт 🔟 строка 1
    for row, l in zip(['L85', 'P85', 'Q85', 'R85', 'S85', 'T85', 'U85', 'V85'], [3, 4, 5, 6, 7, 8, 9, 10]):
        ws[row] = P_10.iloc[0, l]  
#  заносим данные пункт 10 строка 2
    for row, l in zip(['L86', 'P86', 'Q86', 'R86', 'S86', 'T86', 'U86', 'V86'], [3, 4, 5, 6, 7, 8, 9, 10]):
        ws[row] = P_10.iloc[1, l]
#  заносим данные пункт 10 строка 3    
    for row, l in zip(['L87', 'P87', 'Q87', 'R87', 'S87', 'T87', 'U87', 'V87'], [3, 4, 5, 6, 7, 8, 9, 10]):
        ws[row] = P_10.iloc[2, l]  
#  заносим данные пункт 10 строка 4    
    for row, l in zip(['L88', 'P88', 'Q88', 'R88', 'S88', 'T88', 'U88', 'V88'], [3, 4, 5, 6, 7, 8, 9, 10]):
        ws[row] = P_10.iloc[3, l]     
#  заносим данные пункт 10 строка 5
    for row, l in zip(['L89', 'P89', 'Q89', 'R89', 'S89', 'T89', 'U89', 'V89'], [3, 4, 5, 6, 7, 8, 9, 10]):
        ws[row] = P_10.iloc[4, l] 

             
        
    try:
        img = op.drawing.image.Image(all_files[0][0:-6] + i + '.PNG')
        img.anchor = 'A83'
        img.height = 934
        img.width = 1486
        ws.add_image(img)    
    except:
        pass
    
    #img = op.drawing.image.Image(all_files_2[q])
    #img = op.drawing.image.Image(all_files_2[w])
    #img = op.drawing.image.Image(all_files_2[e])
    try:
        img = op.drawing.image.Image(all_files_2[0][0:-8] + i + '-1.PNG')
        img.anchor = 'L93'
        img.height = 550
        img.width = 450
        ws.add_image(img)
    except:
        pass  
        
    try:
        img = op.drawing.image.Image(all_files_2[0][0:-8] + i + '-2.PNG')
        img.anchor = 'P93'
        img.height = 550
        img.width = 450
        ws.add_image(img)
    except:
        pass
    
    try:
        img = op.drawing.image.Image(all_files_2[0][0:-8] + i + '-3.PNG')
        img.anchor = 'T93'
        img.height = 550
        img.width = 380
        ws.add_image(img)
    except:
        pass
    
    #ws['A5'].fill = PatternFill(fill_type='solid', start_color = 'ff8327')  
    #print(ws)    
    #j += 1
    #q += 3
    #w += 3
    #e += 3
    

    wb.save('Z:\\8-Отдел анализа данных\\001 - Сотрудники\\05 - Чурахин А.А\\Долговая карта Python\\'+i+'.xlsx')
    


# # 🅱 Разбивка КН

# In[ ]:


data_history.tail(11)


# In[ ]:


data_general_7['9970'] = data_general_7.pop('КН10')
data_general_7['9971'] = data_general_7.pop('КН01')
data_general_7['9972'] = data_general_7.pop('КН02')
data_general_7['9973'] = data_general_7.pop('КН03')
data_general_7['9974'] = data_general_7.pop('КН04')
data_general_7['9975'] = data_general_7.pop('КН05')
data_general_7['9976'] = data_general_7.pop('КН06')
data_general_7['9977'] = data_general_7.pop('КН07')
#data_general_7['9978'] = data_general_7.pop('КН08')
data_general_7['9979'] = data_general_7.pop('КН09')


# In[ ]:


data_22_KN = ['9970', '9971', '9972', '9973', '9974', '9975', '9976', '9977', '9978', '9979']


# In[ ]:


all_files[0][0:90]


# In[ ]:


#j = 84

for i in data_22_KN:
    #j = 0
    # пункт 1️⃣🅱
    P_1_history = data_history.query('Unnamed == "RF" | Unnamed == @i') # пункт 1 история
    P_1_new= data_new.query('Unnamed == "RF" | Unnamed == @i')          # пункт 1 новое
    P_1_history = P_1_history.drop(['Unnamed'], axis = 1)
    P_1_new = P_1_new.drop(['Unnamed'], axis = 1)
    P_1_new.rename(columns = {'Unnamed: 1' : ''}, inplace = True)
    P_1_history = P_1_history.sort_index(ascending=True)
    P_1_new = P_1_new.sort_index(ascending=True)    
    
    
    # пункт 3️⃣🅱
    P_3 = data_3.query('numb == "RF" | numb == "BK" | numb == @i') # пункт 3
    P_3 = P_3.sort_index(ascending=False)      
    
    # пункт 6️⃣🅱
    P_6_history = data_history_6.query('reg == "RF" | reg == @i') # пункт 6 история
    P_6_new= data_new_6.query('reg == "RF" | reg == @i')          # пункт 6 новое
    P_6_history = P_6_history.sort_index(ascending=True)
    P_6_new = P_6_new.sort_index(ascending=True)
    
    # пункт 7️⃣🅱
    try:
        P_7 = data_general_7[i]
    except:
        pass
    
    # пункт 9️⃣🅱
    P_9 = data_9.query('reg == "Российская Федерация" | reg == @i') # пункт 9
    P_9 = P_9.sort_index(ascending=True)  
    
    # пункт 🔟🅱
    P_10 = data_10_KN_general.query('numb == @i')
    
# НАЧАЛЬНИК
    P_nachalnik = nachalnik_data.query('reg == @i') 
    # ЗАМЫ
    P_zami = zami_data.query('reg == @i')
    
    
    shutil.copy(r"Z:\8-Отдел анализа данных\001 - Сотрудники\05 - Чурахин А.А\Долговая карта Python\Шаблон КН новый.xlsx", 
                i+'.xlsx')
    wb = op.load_workbook(i+'.xlsx')    
    ws = wb.active
    
    
    
###  ШАПКА ОБЩАЯ НАЗВАНИЕ KN
    if i == '9970':
        ws['A2'] = 'Межрегиональной инспекции ФНС России по крупнейшим налогоплательщикам №10'
    elif i == '9971':
        ws['A2'] = 'Межрегиональной инспекции ФНС России по крупнейшим налогоплательщикам №1'
    elif i == '9972':
        ws['A2'] = 'Межрегиональной инспекции ФНС России по крупнейшим налогоплательщикам №2'
    elif i == '9973':
        ws['A2'] = 'Межрегиональной инспекции ФНС России по крупнейшим налогоплательщикам №3'
    elif i == '9974':
        ws['A2'] = 'Межрегиональной инспекции ФНС России по крупнейшим налогоплательщикам №4'
    elif i == '9975':
        ws['A2'] = 'Межрегиональной инспекции ФНС России по крупнейшим налогоплательщикам №5'     
    elif i == '9976':
        ws['A2'] = 'Межрегиональной инспекции ФНС России по крупнейшим налогоплательщикам №6'
    elif i == '9977':
        ws['A2'] = 'Межрегиональной инспекции ФНС России по крупнейшим налогоплательщикам №7'
    elif i == '9978':
        ws['A2'] = 'Межрегиональной инспекции ФНС России по крупнейшим налогоплательщикам №8'    
    elif i == '9979':
        ws['A2'] = 'Межрегиональной инспекции ФНС России по крупнейшим налогоплательщикам №9' 
        
        

#  заносим НАЧАЛЬНИК
    try:
        ws['M1'] = P_nachalnik.iloc[0, 2]             # ФИО
        ws['M2'] = P_nachalnik.iloc[0, 3]             # телефон
    except:
        pass    
    
#  заносим ЗАМЫ
    try:
        for row, l in zip(['S1', 'R2', 'S2'], [4, 6, 5]):
            ws[row] = P_zami.iloc[0, l]
    except:
        pass
    
    
    
# заносим шапку пункт 1️⃣🅱
    ws['C6'] = head_c11
    ws['G6'] = head_g11
    ws['K6'] = head_k11 + ' на ' + head_c11[2:] + ' к сведению на ' + head_g11[2:]
    ws['C11'] = head_c16
    ws['G11'] = head_g16
    ws['K11'] = head_k16 + ' на ' + head_c16[2:] + ' к сведению на ' + head_g16[2:]  
    
  #  заносим данные пункт 1 строчка 1
    for row, l in zip(['A9', 'C9', 'D9', 'E9', 'F9', 'G9', 'H9', 'I9', 'J9', 'K9', 'L9', 'M9', 'N9', 'O9', 
                       'P9', 'Q9', 'R9' ], range(0, 17)):
        ws[row] = P_1_history.iloc[0, l]  
  #  заносим данные пункт 1 строчка 2   
    for row, l in zip(['A10', 'C10', 'D10', 'E10', 'F10', 'G10', 'H10', 'I10', 'J10', 'K10', 'L10', 'M10', 
                       'N10', 'O10', 'P10', 'Q10', 'R10' ], range(0, 17)):
        ws[row] = P_1_history.iloc[1, l]
# закрашивание        
    for row, l in zip(['L10', 'N10', 'P10', 'R10' ], range(10, 17, 2)):
        if P_1_history.iloc[0, l] > P_1_history.iloc[1, l]:
            ws[row].fill = PatternFill(fill_type='solid', start_color = '54bd55')
        else: ws[row].fill = PatternFill(fill_type='solid', start_color = 'ed2213') 
        
  # заносим данные пункт 1 строчка 3      
    for row, l in zip(['A12', 'C12', 'D12', 'E12', 'F12', 'G12', 'H12', 'I12', 'J12', 'K12', 'L12', 'M12', 
                       'N12', 'O12', 'P12', 'Q12', 'R12' ], range(0, 17)):
        ws[row] = P_1_new.iloc[0, l] 
  # заносим данные пункт 1 строчка 4  
    for row, l in zip(['A13', 'C13', 'D13', 'E13', 'F13', 'G13', 'H13', 'I13', 'J13', 'K13', 'L13', 'M13', 
                       'N13', 'O13', 'P13', 'Q13', 'R13' ], range(0, 17)):
        ws[row] = P_1_new.iloc[1, l]
# закрашивание        
    for row, l in zip(['L13', 'N13', 'P13', 'R13' ], range(10, 17, 2)):
        if P_1_new.iloc[0, l] > P_1_new.iloc[1, l]:
            ws[row].fill = PatternFill(fill_type='solid', start_color = '54bd55')
        else: ws[row].fill = PatternFill(fill_type='solid', start_color = 'ed2213')   
  
    
    
   
   #  заносим шапку пункт 3️⃣🅱
    ws['C29'] = head_c30
    ws['F29'] = head_f30    
   #  заносим данные пункт 3 строчка 1
    for row, l in zip(['A31', 'C31', 'D31', 'E31', 'F31', 'G31', 'H31'], range(1, 8)):
        ws[row] = P_3.iloc[0, l]     
   #  заносим данные пункт 3 строчка 2
    for row, l in zip(['A32', 'C32', 'D32', 'E32', 'F32', 'G32', 'H32'], range(1, 8)):
        ws[row] = P_3.iloc[1, l]    
  #  заносим данные пункт 3 строчка 3      
    for row, l in zip(['A33', 'C33', 'D33', 'E33', 'F33', 'G33', 'H33'], range(1, 8)):
        ws[row] = P_3.iloc[2, l] 
        
# закрашивание         
    for row, l in zip(['E33', 'H33' ], [4, 7]):
        if P_3.iloc[0, l] > P_3.iloc[2, l]:
            ws[row].fill = PatternFill(fill_type='solid', start_color = '54bd55')
        else: ws[row].fill = PatternFill(fill_type='solid', start_color = 'ed2213') 
            
    

    
    #  заносим шапку пункт 6️⃣ 🅱   
    ws['G16'] = 'на 01.01.2022 г.' #head_a55  в шапке указан январь, а не номер месяца 
    ws['G22'] = head_a61 
 #  заносим данные пункт 6 строчка 1
    try:
        for row, l in zip(['G20', 'I20', 'J20', 'K20', 'L20', 'M20', 'N20', 'O20', 'P20', 'Q20', 
                       'R20', 'S20', 'T20', 'U20', 'V20', 'W20'], 
                      [1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 11, 12, 13, 14, 15, 16]):
            ws[row] = P_6_history.iloc[0, l] 
    except:
        0
  #  заносим данные пункт 6 строчка 2
    try:
        for row, l in zip(['G21', 'I21', 'J21', 'K21', 'L21', 'M21', 'N21', 'O21', 'P21', 'Q21', 
                       'R21', 'S21', 'T21', 'U21', 'V21', 'W21'], 
                      [1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 11, 12, 13, 14, 15, 16]):
            ws[row] = P_6_history.iloc[1, l]
    except:
        0        
  # заносим данные пункт 6 строчка 3   
    try:
        for row, l in zip(['G23', 'I23', 'J23', 'K23', 'L23', 'M23', 'N23', 'O23', 'P23', 'Q23', 
                       'R23', 'S23', 'T23', 'U23', 'V23', 'W23'], 
                      [1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 11, 12, 13, 14, 15, 16]):
            ws[row] = P_6_new.iloc[0, l] 
    except:
        0       
  # заносим данные пункт 6 строчка 4   
    try:
        for row, l in zip(['G24', 'I24', 'J24', 'K24', 'L24', 'M24', 'N24', 'O24', 'P24', 'Q24', 
                       'R24', 'S24', 'T24', 'U24', 'V24', 'W24'], 
                      [1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 11, 12, 13, 14, 15, 16,]):
            ws[row] = P_6_new.iloc[1, l] 
    except:
        0
    
# закрашивание
   
    for row, l in zip(['K21', 'N21', 'Q21', 'T21', 'W21'], [4, 7, 10, 13, 16]):
        if P_6_history.iloc[1, l] > P_6_history.iloc[0, l]:
                ws[row].fill = PatternFill(fill_type='solid', start_color = '54bd55')
        else: ws[row].fill = PatternFill(fill_type='solid', start_color = 'ed2213')
            
# закрашивание   

    for row, l in zip(['K24', 'N24', 'Q24', 'T24', 'W24'], [4, 7, 10, 13, 16]):
        if P_6_new.iloc[1, l] > P_6_new.iloc[0, l]:
                ws[row].fill = PatternFill(fill_type='solid', start_color = '54bd55')
        else: ws[row].fill = PatternFill(fill_type='solid', start_color = 'ed2213') 

    
    


     #  заносим шапку пункт 7️⃣ 🅱 
    ws['A15'] = 'Количество налогоплательщиков, имеющих задолженность по налогам, сборам,     страховым взносам, пеням и налоговым санкциям на ' + head_a61 
    #  заносим данные пункт 7 строчка 1
    for row, l in zip(['C19', 'D19'], [2, 3]):
        ws[row] = P_7.iloc[0, l] 
    ws['E19'] = P_7.iloc[0, 4] + P_7.iloc[0, 5]  
  #  заносим данные пункт 7 строчка 2   
    for row, l in zip(['C20', 'D20'], [2, 3]):
        ws[row] = P_7.iloc[1, l]
    ws['E20'] = P_7.iloc[1, 4] + P_7.iloc[1, 5]     
  # заносим данные пункт 7 строчка 3      
    for row, l in zip(['C21', 'D21'], [2, 3]):
        ws[row] = P_7.iloc[2, l] 
    ws['E21'] = P_7.iloc[2, 4] + P_7.iloc[2, 5]     
  # заносим данные пункт 7 строчка 4     
    for row, l in zip(['C22', 'D22'], [2, 3]):
        ws[row] = P_7.iloc[3, l] 
    ws['E22'] = P_7.iloc[3, 4] + P_7.iloc[3, 5]     
# заносим данные пункт 7 строчка 5      
    for row, l in zip(['C23', 'D23'], [2, 3]):
        ws[row] = P_7.iloc[4, l]
    ws['E23'] = P_7.iloc[4, 4] + P_7.iloc[4, 5]     
  # заносим данные пункт 7 строчка 6     
    for row, l in zip(['C24', 'D24'], [2, 3]):
        ws[row] = P_7.iloc[5, l]  
    ws['E24'] = P_7.iloc[5, 4] + P_7.iloc[5, 5]  
                
    
   
    #  заносим шапку пункт 9️⃣ 🅱  
    ws['K29'] = head_g75     
    #  заносим данные пункт 9
    try:
        for row, l in zip(['K32', 'Q32', 'R32', 'S32', 'T32'], [1, 5, 2, 3, 4]):
            ws[row] = P_9.iloc[0, l]
        for row, l in zip(['K33', 'M33', 'Q33', 'R33', 'S33', 'T33'], [0, 1, 5, 2, 3, 4]):
            ws[row] = P_9.iloc[1, l]
    except:        
        for row, l in zip(['Q33', 'R33', 'S33', 'T33'], [5, 2, 3, 4]):
            ws[row] = 0
            
        ws['K33'] = i 
        
        if i == '9970':            
            ws['M33'] = 'МИ ФНС России по КН №10'
        elif i == '9971':
            ws['M33'] = 'МИ ФНС России по КН №1'
        elif i == '9972':
            ws['M33'] = 'МИ ФНС России по КН №2'
        elif i == '9973':
            ws['M33'] = 'МИ ФНС России по КН №3'
        elif i == '9974':
            ws['M33'] = 'МИ ФНС России по КН №4'
        elif i == '9975':
            ws['M33'] = 'МИ ФНС России по КН №5'     
        elif i == '9976':
            ws['M33'] = 'МИ ФНС России по КН №6'
        elif i == '9977':
            ws['M33'] = 'МИ ФНС России по КН №7'
        elif i == '9978':
            ws['M33'] = 'МИ ФНС России по КН №8'    
        elif i == '9979':
            ws['M33'] = 'МИ ФНС России по КН №9'  
            
# закрашивание             
    try:
        for row, l in zip(['Q33', 'R33', 'S33', 'T33'], [5, 2, 3, 4]):
            if P_9.iloc[0, l] <  P_9.iloc[1, l]:
                ws[row].fill = PatternFill(fill_type='solid', start_color = '54bd55')
            else: ws[row].fill = PatternFill(fill_type='solid', start_color = 'ed2213')   
    except:
        pass        
       
    
    
#  заносим данные пункт 🔟 🅱 строка 1
    try:
        for row, l in zip(['B42', 'C42', 'D42', 'E42', 'F42', 'G42', 'H42', 'I42', 'J42',
                      'K42', 'L42', 'M42'], range(1, 13)):
            ws[row] = P_10.iloc[0, l] 
    except:
        pass
    
#  заносим данные пункт 10 строка 2
    try:
        for row, l in zip(['B43', 'C43', 'D43', 'E43', 'F43', 'G43', 'H43', 'I43', 'J43',
                      'K43', 'L43', 'M43'], range(1, 13)):
            ws[row] = P_10.iloc[1, l]
    except:
        pass
    
#  заносим данные пункт 10 строка 3  
    try:
        for row, l in zip(['B44', 'C44', 'D44', 'E44', 'F44', 'G44', 'H44', 'I44', 'J44',
                      'K44', 'L44', 'M44'], range(1, 13)):
            ws[row] = P_10.iloc[2, l]  
    except:
        pass
    
#  заносим данные пункт 10 строка 4
    try:
        for row, l in zip(['B45', 'C45', 'D45', 'E45', 'F45', 'G45', 'H45', 'I45', 'J45',
                      'K45', 'L45', 'M45'], range(1, 13)):
            ws[row] = P_10.iloc[3, l]
    except:
        pass
    
#  заносим данные пункт 10 строка 5
    try:
        for row, l in zip(['B46', 'C46', 'D46', 'E46', 'F46', 'G46', 'H46', 'I46', 'J46',
                      'K46', 'L46', 'M46'], range(1, 13)):
            ws[row] = P_10.iloc[4, l]
    except:
        pass
                    
       
    try:
        img = op.drawing.image.Image(all_files[0][0:-6] + i + '.PNG')
        img.anchor = 'A49'
        img.height = 934
        img.width = 1486
        ws.add_image(img)   
    except:
        pass
    
    #j += 1
    
    

    wb.save('Z:\\8-Отдел анализа данных\\001 - Сотрудники\\05 - Чурахин А.А\\Долговая карта Python\\'+i+'.xlsx')
    
    


# In[ ]:




